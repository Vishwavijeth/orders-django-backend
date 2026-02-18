import os
from django.conf import settings
from django.db import transaction
from openpyxl import Workbook, load_workbook

from .models import Order, ReportCheckpoint


REPORT_NAME = "orders_report"
FILE_NAME = "order_details_report.xlsx"


class OrderReportService:

    @staticmethod
    def _get_file_path():
        os.makedirs(settings.MEDIA_ROOT, exist_ok=True)
        return os.path.join(settings.MEDIA_ROOT, FILE_NAME)

    @staticmethod
    def _create_file_with_header(path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders Report"

        ws.append([
            "S.No",
            "User ID",
            "Username",
            "Ordered Items",
            "Restaurant Name",
            "Payment Status",
            "Payment ID",
            "Total Price",
        ])

        wb.save(path)

    @classmethod
    def generate_report(cls):
        """
        idempotent
        """

        file_path = cls._get_file_path()
        file_exists = os.path.exists(file_path)

        # Create file if missing
        if not file_exists:
            cls._create_file_with_header(file_path)

        with transaction.atomic():
            checkpoint, _ = (
                ReportCheckpoint.objects
                .select_for_update()
                .get_or_create(report_name=REPORT_NAME)
            )

            if not file_exists:
                checkpoint.last_order_id = 0
                checkpoint.save(update_fields=["last_order_id"])

            last_id = checkpoint.last_order_id

            orders = (
                Order.objects
                .filter(id__gt=last_id)
                .select_related("user", "restaurant", "payment")
                .prefetch_related("items")
                .order_by("id")
            )

            if not orders.exists():
                return file_path, 0

            wb = load_workbook(file_path)
            ws = wb.active

            serial = ws.max_row  # continues serial numbering
            max_processed_id = last_id
            new_count = 0

            for order in orders.iterator(chunk_size=2000):

                ordered_items = ", ".join(
                    f"{item.menu_name_snapshot} x{item.quantity}"
                    for item in order.items.all()
                )

                ws.append([
                    serial,
                    order.user.id if order.user else None,
                    order.user.username if order.user else None,
                    ordered_items,
                    order.restaurant.name if order.restaurant else None,
                    order.payment.status if order.payment else None,
                    order.payment.payment_id if order.payment else None,
                    order.total_amount,
                ])

                serial += 1
                max_processed_id = order.id
                new_count += 1

            wb.save(file_path)

            checkpoint.last_order_id = max_processed_id
            checkpoint.save(update_fields=["last_order_id"])

        return file_path, new_count
